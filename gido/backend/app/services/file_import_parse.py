# Copyright 2026 玑渡 GIDO Contributors
# SPDX-License-Identifier: Apache-2.0
"""本地文件导入：CSV/Excel 解析与类型推断。"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

LOGICAL_TYPES = ("string", "bigint", "double", "boolean", "datetime")

_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_BOOL_TRUE = {"true", "1", "yes", "y", "是", "t"}
_BOOL_FALSE = {"false", "0", "no", "n", "否", "f"}
_DT_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%d",
    "%Y/%m/%d",
)


def sanitize_column_name(raw: str, used: set[str], idx: int) -> str:
    name = (raw or "").strip()
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"[^\w\u4e00-\u9fff]", "_", name, flags=re.UNICODE)
    name = re.sub(r"_+", "_", name).strip("_")
    if not name:
        name = f"col_{idx + 1}"
    if name[0].isdigit():
        name = f"c_{name}"
    if not _IDENT_RE.match(name) and not re.match(r"^[\w\u4e00-\u9fff]+$", name):
        name = f"col_{idx + 1}"
    base = name[:64]
    cand = base
    n = 2
    while cand.lower() in used:
        cand = f"{base}_{n}"[:64]
        n += 1
    used.add(cand.lower())
    return cand


def _try_parse_datetime(s: str) -> bool:
    text = s.strip()
    if not text:
        return False
    for fmt in _DT_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def infer_cell_type(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "bigint"
    if isinstance(value, float):
        return "double"
    if isinstance(value, datetime):
        return "datetime"
    s = str(value).strip()
    if s == "":
        return None
    low = s.lower()
    if low in _BOOL_TRUE or low in _BOOL_FALSE:
        return "boolean"
    if re.fullmatch(r"[+-]?\d+", s):
        return "bigint"
    if re.fullmatch(r"[+-]?(\d+\.\d*|\.\d+)([eE][+-]?\d+)?", s) or re.fullmatch(
        r"[+-]?\d+[eE][+-]?\d+", s
    ):
        return "double"
    if _try_parse_datetime(s):
        return "datetime"
    return "string"


_TYPE_RANK = {"boolean": 1, "bigint": 2, "double": 3, "datetime": 4, "string": 5}


def merge_types(a: Optional[str], b: Optional[str]) -> Optional[str]:
    if a is None:
        return b
    if b is None:
        return a
    if a == b:
        return a
    # bigint + double → double；其余冲突升 string
    if {a, b} == {"bigint", "double"}:
        return "double"
    if _TYPE_RANK.get(a, 5) == _TYPE_RANK.get(b, 5):
        return "string"
    return a if _TYPE_RANK.get(a, 5) > _TYPE_RANK.get(b, 5) else b


def infer_columns(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> List[Dict[str, Any]]:
    used: set[str] = set()
    names = [sanitize_column_name(h, used, i) for i, h in enumerate(headers)]
    types: List[Optional[str]] = [None] * len(names)
    for row in rows:
        for i in range(len(names)):
            cell = row[i] if i < len(row) else None
            types[i] = merge_types(types[i], infer_cell_type(cell))
    out: List[Dict[str, Any]] = []
    for i, name in enumerate(names):
        out.append(
            {
                "name": name,
                "type": types[i] or "string",
                "nullable": True,
                "is_primary_key": False,
                "source_header": str(headers[i]) if i < len(headers) else name,
            }
        )
    return out


def guess_delimiter(sample: str) -> str:
    candidates = [",", "\t", ";", "|"]
    best = ","
    best_score = -1
    for d in candidates:
        try:
            reader = csv.reader(io.StringIO(sample), delimiter=d)
            rows = []
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= 4:
                    break
            if len(rows) < 2:
                continue
            widths = [len(r) for r in rows if r]
            if not widths or max(widths) < 2:
                continue
            # 列数稳定且 >1 得分高
            if len(set(widths)) == 1:
                score = widths[0] * 10
            else:
                score = max(widths)
            if score > best_score:
                best_score = score
                best = d
        except Exception:
            continue
    return best


def decode_text(raw: bytes, encoding: Optional[str] = None) -> Tuple[str, str]:
    tried = []
    if encoding:
        tried.append(encoding)
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        if enc not in tried:
            tried.append(enc)
    last_err: Optional[Exception] = None
    for enc in tried:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise ValueError(f"无法解码文件编码（尝试 {', '.join(tried)}）: {last_err}")


def parse_csv_bytes(
    raw: bytes,
    *,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    max_rows: int = 500_000,
    preview_rows: int = 100,
) -> Dict[str, Any]:
    text, used_enc = decode_text(raw, encoding)
    sample = text[:8192]
    delim = delimiter if delimiter is not None and delimiter != "" else guess_delimiter(sample)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    all_rows: List[List[str]] = []
    for row in reader:
        all_rows.append(row)
        if len(all_rows) > max_rows + 1:
            raise ValueError(f"行数超过上限 {max_rows}")
    if not all_rows:
        raise ValueError("文件无有效数据行")

    if has_header:
        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        data_rows = all_rows[1:]
    else:
        width = max(len(r) for r in all_rows)
        headers = [f"col_{i + 1}" for i in range(width)]
        data_rows = all_rows

    # 对齐列宽
    width = len(headers)
    norm_rows: List[List[Any]] = []
    for r in data_rows:
        cells = list(r[:width]) + [""] * max(0, width - len(r))
        norm_rows.append(cells)

    columns = infer_columns(headers, norm_rows[: min(len(norm_rows), 2000)])
    preview = norm_rows[:preview_rows]
    return {
        "format": "csv",
        "encoding": used_enc,
        "delimiter": delim,
        "has_header": has_header,
        "columns": columns,
        "preview_rows": preview,
        "all_rows": norm_rows,
        "row_count": len(norm_rows),
        "truncated": False,
    }


def parse_xlsx_bytes(
    raw: bytes,
    *,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 500_000,
    preview_rows: int = 100,
) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("服务端未安装 openpyxl，无法解析 Excel") from e

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            ws = wb[wb.sheetnames[0]]
            used_sheet = wb.sheetnames[0]

        all_rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
            if len(all_rows) > max_rows + 1:
                raise ValueError(f"行数超过上限 {max_rows}")
    finally:
        wb.close()

    if not all_rows:
        raise ValueError("Excel 工作表无有效数据")

    # 去掉全空行
    all_rows = [r for r in all_rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not all_rows:
        raise ValueError("Excel 工作表无有效数据")

    if has_header:
        headers = ["" if c is None else str(c) for c in all_rows[0]]
        data_rows = all_rows[1:]
    else:
        width = max(len(r) for r in all_rows)
        headers = [f"col_{i + 1}" for i in range(width)]
        data_rows = all_rows

    width = len(headers)
    norm_rows: List[List[Any]] = []
    for r in data_rows:
        cells = list(r[:width]) + [None] * max(0, width - len(r))
        # 统一成可预览字符串 / 原值
        norm_rows.append(["" if c is None else c for c in cells])

    columns = infer_columns(headers, norm_rows[: min(len(norm_rows), 2000)])
    preview = [
        ["" if c is None else (c.isoformat(sep=" ") if isinstance(c, datetime) else c) for c in row]
        for row in norm_rows[:preview_rows]
    ]
    return {
        "format": "xlsx",
        "encoding": "binary",
        "delimiter": None,
        "has_header": has_header,
        "sheet_name": used_sheet,
        "sheet_names": [],  # filled by caller if needed
        "columns": columns,
        "preview_rows": preview,
        "all_rows": norm_rows,
        "row_count": len(norm_rows),
        "truncated": False,
    }


def list_xlsx_sheets(raw: bytes) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def parse_file_bytes(
    raw: bytes,
    fmt: str,
    *,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 500_000,
    preview_rows: int = 100,
) -> Dict[str, Any]:
    fmt = (fmt or "").lower()
    if fmt == "csv":
        return parse_csv_bytes(
            raw,
            encoding=encoding,
            delimiter=delimiter,
            has_header=has_header,
            max_rows=max_rows,
            preview_rows=preview_rows,
        )
    if fmt == "xlsx":
        result = parse_xlsx_bytes(
            raw,
            has_header=has_header,
            sheet_name=sheet_name,
            max_rows=max_rows,
            preview_rows=preview_rows,
        )
        try:
            result["sheet_names"] = list_xlsx_sheets(raw)
        except Exception:
            result["sheet_names"] = [result.get("sheet_name")] if result.get("sheet_name") else []
        return result
    raise ValueError(f"不支持的文件格式: {fmt}")


def coerce_cell(value: Any, logical_type: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    t = (logical_type or "string").lower()
    if t == "string":
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        return str(value)
    if t == "bigint":
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        s = str(value).strip()
        return int(float(s)) if re.search(r"[.eE]", s) else int(s)
    if t == "double":
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        return float(str(value).strip())
    if t == "boolean":
        if isinstance(value, bool):
            return 1 if value else 0
        if isinstance(value, (int, float)):
            return 1 if value else 0
        low = str(value).strip().lower()
        if low in _BOOL_TRUE:
            return 1
        if low in _BOOL_FALSE:
            return 0
        raise ValueError(f"无法解析为 boolean: {value!r}")
    if t == "datetime":
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        s = str(value).strip()
        for fmt in _DT_FORMATS:
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        raise ValueError(f"无法解析为 datetime: {value!r}")
    return str(value)


def load_all_rows(
    raw: bytes,
    fmt: str,
    *,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 500_000,
) -> Dict[str, Any]:
    """装载用（小文件）：返回 columns 元信息 + 全部数据行。大文件请用 iter_*_rows_from_path。"""
    return parse_file_bytes(
        raw,
        fmt,
        encoding=encoding,
        delimiter=delimiter,
        has_header=has_header,
        sheet_name=sheet_name,
        max_rows=max_rows,
        preview_rows=min(100, max_rows),
    )


def _sniff_csv_encoding(path: "Path", encoding: Optional[str] = None) -> str:
    from pathlib import Path

    p = Path(path)
    head = p.read_bytes()[:65536]
    _, used = decode_text(head, encoding)
    return used


def parse_csv_path(
    path: "Path",
    *,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    max_rows: int = 5_000_000,
    preview_rows: int = 100,
    infer_rows: int = 2000,
) -> Dict[str, Any]:
    """大文件友好：只抽样推断类型与预览，行数可估算。"""
    from pathlib import Path

    p = Path(path)
    size = p.stat().st_size
    used_enc = _sniff_csv_encoding(p, encoding)
    with p.open("rb") as bf:
        sample_bytes = bf.read(65536)
    sample_text, _ = decode_text(sample_bytes, used_enc)
    delim = delimiter if delimiter is not None and delimiter != "" else guess_delimiter(sample_text)

    infer_budget = max(infer_rows, preview_rows) + (1 if has_header else 0)
    collected: List[List[str]] = []
    exact_count = 0

    with p.open("r", encoding=used_enc, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delim)
        for row in reader:
            exact_count += 1
            if len(collected) < infer_budget:
                collected.append(row)
            if exact_count > max_rows + (1 if has_header else 0):
                raise ValueError(f"行数超过上限 {max_rows}")

    if not collected:
        raise ValueError("文件无有效数据行")

    if has_header:
        headers = [str(c) if c is not None else "" for c in collected[0]]
        sample_data = collected[1:]
        header_offset = 1
    else:
        width = max(len(r) for r in collected)
        headers = [f"col_{i + 1}" for i in range(width)]
        sample_data = collected
        header_offset = 0

    width = len(headers)
    norm_sample: List[List[Any]] = []
    for r in sample_data:
        cells = list(r[:width]) + [""] * max(0, width - len(r))
        norm_sample.append(cells)

    columns = infer_columns(headers, norm_sample[:infer_rows])
    preview = norm_sample[:preview_rows]
    row_count = max(0, exact_count - header_offset)

    return {
        "format": "csv",
        "encoding": used_enc,
        "delimiter": delim,
        "has_header": has_header,
        "columns": columns,
        "preview_rows": preview,
        "row_count": row_count,
        "row_count_estimated": False,
        "truncated": False,
        "size_bytes": size,
    }


def parse_xlsx_path(
    path: "Path",
    *,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 5_000_000,
    preview_rows: int = 100,
    infer_rows: int = 2000,
) -> Dict[str, Any]:
    from pathlib import Path
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            ws = wb[sheet_names[0]]
            used_sheet = sheet_names[0]

        collected: List[List[Any]] = []
        exact = 0
        for row in ws.iter_rows(values_only=True):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            exact += 1
            if exact > max_rows + (1 if has_header else 0):
                raise ValueError(f"行数超过上限 {max_rows}")
            if len(collected) < max(infer_rows, preview_rows) + (1 if has_header else 0):
                collected.append(list(row))
            # Excel 单 sheet 理论 ≤1048576；继续计数到完
        # 若只采了样，再扫一遍计数代价高；read_only 已顺序读完则 exact 是全量
        # 上面循环已读完全表（因为没有 break），exact 为全量非空行
    finally:
        wb.close()

    if not collected and exact == 0:
        raise ValueError("Excel 工作表无有效数据")

    # 若 exact > collected，需要再读一次取全量计数——上面已全量扫描，但只保留了 sample。
    # 重新打开仅计数代价翻倍。改为在同一循环里只存 sample，exact 全量累加。
    # 当前实现：循环没有在 sample 满后 break，所以 exact 正确，collected 只保留前 N——不对！
    # 我写的是 if len(collected) < ...: append，所以 exact 全量正确。

    if has_header:
        headers = ["" if c is None else str(c) for c in collected[0]]
        sample_data = collected[1:]
        row_count = max(0, exact - 1)
    else:
        width = max((len(r) for r in collected), default=0)
        headers = [f"col_{i + 1}" for i in range(width)]
        sample_data = collected
        row_count = exact

    width = len(headers)
    norm_sample: List[List[Any]] = []
    for r in sample_data:
        cells = list(r[:width]) + [None] * max(0, width - len(r))
        norm_sample.append(["" if c is None else c for c in cells])

    columns = infer_columns(headers, norm_sample[:infer_rows])
    preview = [
        ["" if c is None else (c.isoformat(sep=" ") if isinstance(c, datetime) else c) for c in row]
        for row in norm_sample[:preview_rows]
    ]
    return {
        "format": "xlsx",
        "encoding": "binary",
        "delimiter": None,
        "has_header": has_header,
        "sheet_name": used_sheet,
        "sheet_names": sheet_names,
        "columns": columns,
        "preview_rows": preview,
        "row_count": row_count,
        "row_count_estimated": False,
        "truncated": False,
        "size_bytes": p.stat().st_size,
    }


def parse_file_path(
    path: "Path",
    fmt: str,
    *,
    encoding: Optional[str] = None,
    delimiter: Optional[str] = None,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 5_000_000,
    preview_rows: int = 100,
    infer_rows: int = 2000,
) -> Dict[str, Any]:
    fmt = (fmt or "").lower()
    if fmt == "csv":
        return parse_csv_path(
            path,
            encoding=encoding,
            delimiter=delimiter,
            has_header=has_header,
            max_rows=max_rows,
            preview_rows=preview_rows,
            infer_rows=infer_rows,
        )
    if fmt == "xlsx":
        return parse_xlsx_path(
            path,
            has_header=has_header,
            sheet_name=sheet_name,
            max_rows=max_rows,
            preview_rows=preview_rows,
            infer_rows=infer_rows,
        )
    raise ValueError(f"不支持的文件格式: {fmt}")


def iter_csv_rows_from_path(
    path: "Path",
    *,
    encoding: Optional[str] = None,
    delimiter: str = ",",
    has_header: bool = True,
    max_rows: int = 5_000_000,
):
    """逐行产出数据行（不含表头）。"""
    from pathlib import Path

    p = Path(path)
    used_enc = encoding or _sniff_csv_encoding(p, encoding)
    with p.open("r", encoding=used_enc, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter or ",")
        first = True
        n = 0
        for row in reader:
            if first and has_header:
                first = False
                continue
            first = False
            n += 1
            if n > max_rows:
                raise ValueError(f"行数超过上限 {max_rows}")
            yield row


def iter_xlsx_rows_from_path(
    path: "Path",
    *,
    has_header: bool = True,
    sheet_name: Optional[str] = None,
    max_rows: int = 5_000_000,
):
    from pathlib import Path
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        ws = wb[sheet_name] if sheet_name and sheet_name in names else wb[names[0]]
        first = True
        n = 0
        for row in ws.iter_rows(values_only=True):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            if first and has_header:
                first = False
                continue
            first = False
            n += 1
            if n > max_rows:
                raise ValueError(f"行数超过上限 {max_rows}")
            yield ["" if c is None else c for c in row]
    finally:
        wb.close()
